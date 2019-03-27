import requests
from bs4 import BeautifulSoup
import re
import openpyxl

def Make_Soup(start):
    headers = {'User-Agent':'Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/72.0.3626.121 Safari/537.36'}
    r = requests.get('https://movie.douban.com/top250?start='+str(start)+'&filter=', headers = headers)
    soup = BeautifulSoup(r.text, 'html.parser')
    return soup

def Get_Ranks(soup):
    # 排名
    for each in soup.find_all('em'):
        ranks.append(each.string)

def Get_Titles(soup):
    # 标题
    for each in soup.find_all('div','hd'):
        each = each.find('span',class_ = 'title')
        titles.append(each.string)

def Get_Others(soup):
    # 导演 演员 发行年份 发行国 类型 得分 引用
    for each in soup.find_all('div', class_='info'):
        each_ = each.find('div', class_='bd')

        # 包含导演 演员 发行年份 发行国 类型的tag
        each = each_.find('p')
        string = re.split(r'[\n]|：', each.text)
        string1 = re.split(r'\s{2,}|[:]', string[1])
        directors.append(string1[2])
        try:
            actors.append(string1[4])
        except IndexError:
            # 简介中不一定包含主演
            actors.append(' ')
        string2 = re.split(r'/', string[2])
        years.append(re.sub(r'\s+','',string2[0]))
        countries.append(re.sub(r'\s{2,}','',string2[1]))
        types.append(re.sub(r'\s{2,}','',string2[2]))

        # 得分tag
        each = each_.find('span', class_='rating_num')
        scores.append(each.text)

        # 引用tag
        each = each_.find('span', class_='inq')
        # 某些影片没有引用
        try:
            quotes.append('"'+each.string+'"')
        except AttributeError:
            quotes.append('""')


def WriteToTXT():
    with open('豆瓣250.txt','wt',encoding = 'utf-8') as f:
        for i in range(len(titles)):
            f.write(ranks[i] + ' '+ titles[i] + '\n导演：'+ directors[i] + ' 演员：'+ actors[i] + '\n'+ years[i] + ' '+ countries[i]
                    + ' '+ types[i] + ' '+ scores[i] + ' \n'+ quotes[i] + "\n\n")

def WriteToExcel():
    wb = openpyxl.Workbook()
    ws = wb.active

    ws['A1'] = "排名"
    ws['B1'] = "评分"
    ws['C1'] = "电影名称"
    ws['D1'] = "导演"
    ws['E1'] = "演员"
    ws['F1'] = "发行时间"
    ws['G1'] = "发行国"
    ws['H1'] = "类型"
    ws['I1'] = "引用"

    for i in range(len(titles)):
        ws.append([ranks[i], scores[i], titles[i], directors[i], actors[i], years[i], countries[i], types[i], quotes[i]])

    wb.save("豆瓣250.xlsx")


if __name__ == '__main__':

    ranks = []  # 排名
    titles = []  # 标题
    directors = []  # 导演
    actors = []  # 主演
    years = []  # 发行年份
    countries = []  # 发行国
    types = []  # 类型
    scores = []  # 得分
    quotes = []  # 引用

    start = 0
    while start < 250:
        try:
            soup = Make_Soup(start)
        except requests.exceptions.ConnectionError:
            break
        Get_Ranks(soup)
        Get_Titles(soup)
        Get_Others(soup)
        # WriteToTXT()
        WriteToExcel()
        start += 25