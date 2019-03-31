import openpyxl as op

if __name__ == '__main__':
    NeedChange = {'阿甘正传':9.8, '这个杀手不太冷':9.6, '肖申克的救赎':9.7}
    wb = op.load_workbook(r'豆瓣250.xlsx')
    ws = wb[wb.sheetnames[0]]
    rows = str(len(list(ws.rows)))
    finalone = 'C'+ rows
    for each in ws['C0': finalone]:
        if each.value in NeedChange:
            #c = op.cell.cell.column_index_from_string(each_row.column)
            #c = op.cell.cell.get_column_letter(c-1)
            each_score = each.offset(0, -1)
            each_score.value = NeedChange[each.value]
    wb.save('豆瓣250.xlsx')
